#!/usr/bin/env python3
"""
NOTAM scraper for https://notams.aim.faa.gov/notamSearch/nsapp.html#/

Automates the FAA NOTAM search workflow using Playwright, captures the exported
Excel workbook, and converts the information into a structured representation.
A lightweight HTML report is also generated to visualise logs and highlight key
NOTAM details.
"""

from __future__ import annotations

import argparse
import datetime as dt
import html
import io
import json
import logging
import os
import sys
import tempfile
from pathlib import Path
from typing import Any, Dict, List, Tuple
from urllib.parse import urljoin

from bs4 import BeautifulSoup
from openpyxl import load_workbook
from playwright.sync_api import Playwright, TimeoutError, sync_playwright
try:
    import xlrd
except ImportError:  # pragma: no cover - safety net for environments without xlrd
    xlrd = None


LOGGER = logging.getLogger("notam_scraper")
LOG_FORMAT = "%(asctime)s | %(levelname)-8s | %(message)s"
_LOGGING_CONFIGURED = False

# FAA NOTAM portal URL
NOTAM_PORTAL_URL = "https://notams.aim.faa.gov/notamSearch/nsapp.html#/"


def _should_launch_headless() -> bool:
    """Determine whether Chromium should run in headless mode."""
    env_value = os.getenv("NOTAM_SCRAPER_HEADLESS")
    if env_value is not None:
        return env_value.strip().lower() not in {"0", "false", "no"}

    # If there is no display (common on Linux servers), fall back to headless.
    if sys.platform.startswith("linux") and not os.getenv("DISPLAY"):
        return True

    # Default to headed mode to satisfy local UX requirements.
    return False


def configure_logging(verbose: bool = False) -> None:
    """Configure module-level logging."""
    global _LOGGING_CONFIGURED
    level = logging.DEBUG if verbose else logging.INFO

    if not _LOGGING_CONFIGURED:
        logging.basicConfig(level=level, format=LOG_FORMAT)
        _LOGGING_CONFIGURED = True

    LOGGER.setLevel(level)


def log_and_capture(messages: List[str], message: str, level: int = logging.INFO) -> None:
    """Log a message and store it for later HTML reporting."""
    LOGGER.log(level, message)
    messages.append(f"{dt.datetime.utcnow().isoformat()}Z | {logging.getLevelName(level)} | {message}")


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dt.datetime, dt.date)):
        return value.isoformat()
    return str(value)


def parse_html_table_notams(html_bytes: bytes) -> List[Dict[str, str]]:
    """Parse NOTAM entries from an HTML table (legacy Excel export)."""
    soup = BeautifulSoup(html_bytes, "html.parser")
    table = soup.find("table")
    if table is None:
        return [{"header": "NOTAM Export", "raw": html_bytes.decode("utf-8", errors="replace")}]

    rows = []
    for tr in table.find_all("tr"):
        cells = [cell.get_text(strip=True) for cell in tr.find_all(["th", "td"])]
        if cells:
            rows.append(cells)

    if not rows:
        return []

    headers = [(cell or f"Column {idx+1}").strip() for idx, cell in enumerate(rows[0])]
    entries: List[Dict[str, str]] = []

    for row in rows[1:]:
        entry: Dict[str, str] = {}
        for idx, value in enumerate(row):
            if idx < len(headers):
                key = headers[idx]
                if key and value:
                    entry[key] = value

        if entry:
            header = entry.get("NOTAM Number") or entry.get("Number") or entry.get("Condition") or "NOTAM Entry"
            entry["header"] = header
            entry["raw"] = "\n".join(f"{k}: {v}" for k, v in entry.items() if k not in {"header"})
            entries.append(entry)

    return entries


def parse_excel_notams(excel_bytes: bytes) -> List[Dict[str, str]]:
    """Parse NOTAM entries from the downloaded Excel workbook."""
    if excel_bytes.startswith(b"\xD0\xCF\x11\xE0"):
        if xlrd is None:
            raise RuntimeError("Received legacy XLS content but xlrd is unavailable.")
        workbook = xlrd.open_workbook(file_contents=excel_bytes)
        sheet = workbook.sheet_by_index(0)
        entries: List[Dict[str, str]] = []
        parsed_rows: List[List[str]] = []
        for row_idx in range(sheet.nrows):
            row_values: List[str] = []
            for col_idx in range(sheet.ncols):
                cell = sheet.cell(row_idx, col_idx)
                value = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        value = xlrd.xldate_as_datetime(value, workbook.datemode)
                    except (ValueError, OverflowError):
                        value = cell.value
                elif cell.ctype == xlrd.XL_CELL_NUMBER and float(value).is_integer():
                    value = int(value)
                row_values.append(_stringify(value).strip())
            parsed_rows.append(row_values)

        metadata_entry: Dict[str, str] | None = None
        header_row: List[str] | None = None

        if parsed_rows:
            meta = parsed_rows[0]
            meta_pairs = []
            if meta and meta[0]:
                meta_pairs.append(("Report", meta[0]))
            if len(meta) > 1 and meta[1]:
                meta_pairs.append(("Filters", meta[1]))
            if meta_pairs:
                metadata_entry = {
                    "header": "NOTAM Export Metadata",
                    "raw": "\n".join(f"{k}: {v}" for k, v in meta_pairs),
                }
                for k, v in meta_pairs:
                    metadata_entry[k] = v

        for row in parsed_rows[1:]:
            if not any(row):
                continue

            first_cell = row[0].lower() if row and row[0] else ""
            if header_row is None and first_cell == "location":
                header_row = [cell or f"Column {idx+1}" for idx, cell in enumerate(row)]
                continue

            if header_row is None:
                continue

            entry: Dict[str, str] = {}
            for idx, value in enumerate(row):
                if idx >= len(header_row):
                    key = f"Column {idx+1}"
                else:
                    key = header_row[idx] or f"Column {idx+1}"
                if value:
                    entry[key] = value

            if entry:
                entry["header"] = (
                    entry.get("NOTAM #/LTA #")
                    or entry.get("NOTAM Number")
                    or entry.get("Number")
                    or entry.get("Location")
                    or "NOTAM Entry"
                )
                entry["raw"] = "\n".join(f"{k}: {v}" for k, v in entry.items() if k not in {"header"})
                entries.append(entry)

        if metadata_entry:
            entries.insert(0, metadata_entry)

        return entries

    if not excel_bytes.startswith(b"PK\x03\x04"):
        return parse_html_table_notams(excel_bytes)

    workbook = load_workbook(filename=io.BytesIO(excel_bytes), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        return []

    headers = [(_stringify(cell).strip() or f"Column {idx+1}") for idx, cell in enumerate(rows[0])]
    entries: List[Dict[str, str]] = []

    for raw_row in rows[1:]:
        if not raw_row or all(cell is None or _stringify(cell).strip() == "" for cell in raw_row):
            continue

        entry: Dict[str, str] = {}
        for idx, cell in enumerate(raw_row):
            key = headers[idx]
            value = _stringify(cell).strip()
            if key and value:
                entry[key] = value

        if not entry:
            continue

        header = entry.get("NOTAM Number") or entry.get("Number") or entry.get("Condition") or "NOTAM Entry"
        entry["header"] = header
        entry["raw"] = "\n".join(f"{k}: {v}" for k, v in entry.items() if k not in {"header"})
        entries.append(entry)

    return entries


def generate_html_report(
    airport_code: str,
    logs: List[str],
    notams: List[Dict[str, str]],
    output_path: Path,
) -> None:
    """Create a simple HTML page containing scraper logs and NOTAM highlights."""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    log_items = "\n".join(f"<li>{log}</li>" for log in logs)
    notam_blocks: List[str] = []

    for entry in notams:
        header = html.escape(entry.get("header", "Unknown NOTAM"))
        detail_items = [(k, v) for k, v in entry.items() if k not in {"header", "raw"}]
        primary_text = entry.get("Condition") or entry.get("Text") or entry.get("Information") or entry.get("raw", "")
        primary_text = html.escape(primary_text)

        if detail_items:
            metadata_rows = "".join(
                f"<tr><th scope='row'>{html.escape(str(key))}</th><td>{html.escape(str(value))}</td></tr>"
                for key, value in detail_items
            )
            metadata_table = f"<table class='notam-meta'><tbody>{metadata_rows}</tbody></table>"
        else:
            metadata_table = "<p class='meta-placeholder'>No supplementary metadata captured.</p>"

        notam_blocks.append(
            f"""
                <section class="notam-card">
                    <header>
                        <h3>{header}</h3>
                    </header>
                    {metadata_table}
                    <div class="notam-main">
                        <h4>Summary</h4>
                        <p>{primary_text}</p>
                    </div>
                    <details>
                        <summary>View raw entry</summary>
                        <pre>{html.escape(entry.get("raw", ""))}</pre>
                    </details>
                </section>
            """
        )

    notam_html = "\n".join(notam_blocks) if notam_blocks else "<p>No NOTAMs found.</p>"

    html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="utf-8" />
    <title>NOTAM Scraper Report - {airport_code}</title>
    <style>
        :root {{
            color-scheme: light dark;
            font-family: "Segoe UI", Arial, sans-serif;
        }}
        body {{
            margin: 2rem;
            background: #f5f7fa;
            color: #1f2933;
        }}
        h1, h2 {{
            margin-bottom: 0.5rem;
        }}
        .log-panel {{
            background: #ffffff;
            border-radius: 0.75rem;
            padding: 1rem 1.5rem;
            box-shadow: 0 6px 18px rgba(31, 41, 51, 0.1);
            margin-bottom: 2rem;
        }}
        .log-panel ul {{
            list-style: none;
            padding: 0;
            margin: 0;
        }}
        .log-panel li {{
            padding: 0.4rem 0;
            border-bottom: 1px solid rgba(31, 41, 51, 0.1);
            font-family: "Fira Code", Consolas, monospace;
        }}
        .log-panel li:last-child {{
            border-bottom: none;
        }}
        .grid {{
            display: grid;
            gap: 1.5rem;
        }}
        .notam-card {{
            background: linear-gradient(135deg, #ffffff 0%, #f8fbff 100%);
            border-left: 6px solid #2563eb;
            border-radius: 1rem;
            box-shadow: 0 10px 24px rgba(37, 99, 235, 0.12);
            padding: 1.5rem;
        }}
        .notam-card h3 {{
            margin-top: 0;
            color: #1d4ed8;
        }}
        .notam-main {{
            background: rgba(37, 99, 235, 0.08);
            border-radius: 0.75rem;
            padding: 1rem 1.2rem;
            margin-top: 1rem;
        }}
        .notam-main h4 {{
            margin: 0 0 0.5rem 0;
            color: #1e40af;
            text-transform: uppercase;
            letter-spacing: 0.08em;
            font-size: 0.85rem;
        }}
        .notam-main p {{
            margin: 0;
            font-weight: 600;
            color: #0f172a;
        }}
        .notam-meta {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 1rem;
        }}
        .notam-meta th, .notam-meta td {{
            text-align: left;
            padding: 0.35rem 0.6rem;
        }}
        .notam-meta th {{
            width: 3rem;
            color: #475569;
            font-weight: 700;
        }}
        .notam-meta tr:nth-child(odd) {{
            background: rgba(148, 163, 184, 0.1);
        }}
        details {{
            margin-top: 1rem;
        }}
        summary {{
            cursor: pointer;
            color: #2563eb;
            font-weight: 600;
        }}
        pre {{
            background: #0f172a;
            color: #e2e8f0;
            padding: 1rem;
            border-radius: 0.75rem;
            overflow-x: auto;
        }}
        .meta-placeholder {{
            margin-top: 0.5rem;
            font-style: italic;
            color: #64748b;
        }}
    </style>
</head>
<body>
    <header>
        <h1>NOTAM Scraper Report</h1>
        <p>Airport: <strong>{airport_code}</strong> &middot; Generated: {dt.datetime.utcnow().isoformat()}Z</p>
    </header>

    <section class="log-panel">
        <h2>Scraper Activity Log</h2>
        <ul>
            {log_items}
        </ul>
    </section>

    <section>
        <h2>Captured NOTAMs</h2>
        <div class="grid">
            {notam_html}
        </div>
    </section>
</body>
</html>
"""

    output_path.write_text(html_content, encoding="utf-8")


def capture_excel_report(page) -> Tuple[bytes, str]:
    """Click the Excel export button and return the downloaded workbook bytes."""
    button_selector = 'div[ng-click="globalScope.excel();"]'
    page.wait_for_selector(button_selector, state="visible", timeout=60000)

    with page.expect_download(timeout=60000) as download_info:
        page.click(button_selector)
    download = download_info.value
    failure = download.failure()
    if failure:
        raise RuntimeError(f"Excel download failed: {failure}")

    download_path = download.path()
    if download_path:
        excel_bytes = Path(download_path).read_bytes()
    else:
        fd, temp_path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        try:
            download.save_as(temp_path)
            excel_bytes = Path(temp_path).read_bytes()
        finally:
            Path(temp_path).unlink(missing_ok=True)

    filename = download.suggested_filename or (Path(download_path).name if download_path else "notams.xlsx")
    if not filename.lower().endswith(".xlsx"):
        filename = f"{filename}.xlsx"

    return excel_bytes, filename


def run_scraper(
    playwright: Playwright, airport_code: str, logs: List[str]
) -> Tuple[List[Dict[str, str]], bytes, str]:
    """Execute the automated NOTAM workflow and return structured NOTAMs and raw Excel bytes."""
    headless = _should_launch_headless()
    launch_args = ["--disable-http2", "--disable-dev-shm-usage", "--disable-gpu"]
    if headless:
        launch_args.extend(
            ["--disable-blink-features=AutomationControlled", "--disable-software-rasterizer"]
        )

    log_and_capture(logs, f"Launching Chromium (headless={headless}).")
    browser = playwright.chromium.launch(headless=headless, args=launch_args)
    context = browser.new_context(
        ignore_https_errors=True,
        accept_downloads=True,
        user_agent=(
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/129.0.0.0 Safari/537.36"
        ),
        viewport={"width": 1280, "height": 720},
    )
    page = context.new_page()

    try:
        disclaimer_url = NOTAM_PORTAL_URL.replace("nsapp.html#/", "disclaimer.html")

        max_attempts = 3
        for attempt in range(1, max_attempts + 1):
            try:
                log_and_capture(logs, f"Opening disclaimer (attempt {attempt}/{max_attempts}): {disclaimer_url}")
                page.goto(disclaimer_url, wait_until="domcontentloaded", timeout=120000)
                break
            except TimeoutError as exc:
                log_and_capture(
                    logs,
                    f"Timeout reaching disclaimer on attempt {attempt}: {exc}",
                    logging.WARNING,
                )
                if attempt == max_attempts:
                    raise
                page.wait_for_timeout(5000)

        log_and_capture(logs, "Acknowledging disclaimer.")
        page.get_by_role("button", name="I've read and understood above statements").click(force=True)
        page.wait_for_url(NOTAM_PORTAL_URL + "*", timeout=120000)
        page.wait_for_selector('input[name="designatorsForLocation"]', timeout=60000)

        log_and_capture(logs, f"Searching NOTAMs for airport: {airport_code}")
        search_input = page.wait_for_selector('input[name="designatorsForLocation"]', state="visible")
        search_input.fill(airport_code)
        page.click('button:has-text("Search")')

        log_and_capture(logs, "Waiting for search results to load.")
        page.wait_for_selector('span[ng-click="selectAllForPrint();"]', state="visible")

        log_and_capture(logs, "Selecting all NOTAM entries for export.")
        page.click('span[ng-click="selectAllForPrint();"]')

        log_and_capture(logs, "Downloading NOTAM Excel workbook.")
        excel_bytes, filename = capture_excel_report(page)
        log_and_capture(logs, f"Captured Excel payload ({len(excel_bytes)} bytes) from {filename}.")

        notams = parse_excel_notams(excel_bytes)

        log_and_capture(logs, f"Parsed {len(notams)} NOTAM entries from Excel.")
        return notams, excel_bytes, "xlsx"

    except TimeoutError as exc:
        log_and_capture(logs, f"Timeout while interacting with the portal: {exc}", logging.ERROR)
        raise
    finally:
        context.close()
        browser.close()


def save_raw_outputs(
    airport_code: str,
    artifact_bytes: bytes,
    notams: List[Dict[str, str]],
    output_dir: Path,
    *,
    artifact_extension: str,
) -> Tuple[Path, Path]:
    """Persist the downloaded Excel workbook and structured JSON output."""
    timestamp = dt.datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
    airport_slug = airport_code.replace(",", "_").replace(" ", "_")
    base_name = f"{airport_slug}_{timestamp}"

    artifact_path = output_dir / f"{base_name}.{artifact_extension}"
    json_path = output_dir / f"{base_name}.json"

    output_dir.mkdir(parents=True, exist_ok=True)

    artifact_path.write_bytes(artifact_bytes)
    json_path.write_text(json.dumps(notams, indent=2), encoding="utf-8")

    return artifact_path, json_path


def scrape_notams(
    airport_code: str,
    output_dir: Path | None = None,
    *,
    log_html: str = "notam_report.html",
    save_artifacts: bool = False,
    verbose: bool = False,
) -> Dict[str, object]:
    """
    High-level helper that orchestrates the NOTAM scraping flow.

    Returns a dictionary containing the airport code, parsed NOTAM entries, log
    messages collected during execution, and optionally the saved artifact
    locations (PDF/JSON/HTML).
    """
    code = airport_code.strip().upper()
    if not code:
        raise ValueError("Airport code must not be empty.")

    configure_logging(verbose=verbose)

    logs: List[str] = []
    artifact_path: Path | None = None
    json_path: Path | None = None
    html_path: Path | None = None

    with sync_playwright() as playwright:
        notams, artifact_bytes, artifact_extension = run_scraper(playwright, code, logs)

    if save_artifacts and output_dir:
        output_dir = Path(output_dir)
        artifact_path, json_path = save_raw_outputs(
            code, artifact_bytes, notams, output_dir, artifact_extension=artifact_extension
        )
        log_and_capture(logs, f"Saved export to {artifact_path}")

        html_path = output_dir / log_html
        log_and_capture(logs, f"Preparing HTML report at {html_path}")
        generate_html_report(code, logs.copy(), notams, html_path)
        log_and_capture(logs, f"HTML report created at {html_path}")
    else:
        log_and_capture(logs, "Artifacts saving skipped.")

    log_and_capture(logs, "Run complete.", logging.INFO)

    return {
        "airport": code,
        "logs": logs,
        "notams": notams,
        "artifact_path": str(artifact_path) if artifact_path else None,
        "json_path": str(json_path) if json_path else None,
        "html_path": str(html_path) if html_path else None,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Scrape NOTAMs for a given airport code via the FAA portal.")
    parser.add_argument("airport", help="Airport or facility code (comma-separated for multiple).")
    parser.add_argument(
        "--output-dir",
        default="notam_reports",
        type=Path,
        help="Directory where PDF, JSON, and HTML reports will be saved.",
    )
    parser.add_argument(
        "--log-html",
        default="notam_report.html",
        help="Filename for the generated HTML log/report (relative to output directory).",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Enable verbose logging.",
    )

    args = parser.parse_args()
    result = scrape_notams(
        args.airport,
        output_dir=args.output_dir,
        log_html=args.log_html,
        save_artifacts=True,
        verbose=args.verbose,
    )

    print(json.dumps({"airport": result["airport"], "notams": result["notams"]}, indent=2))


if __name__ == "__main__":
    main()

